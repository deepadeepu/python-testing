import openpyxl

book = openpyxl.load_workbook(r"C:\Users\DEEPA\PythonTesting\Testing1\PythonSelFramework\test_data.xlsx")
sheet = book.active


class HomePageData:
    # test_HomePage_Data=[{"name":"deepa","email":"deepa","Password":"Ram"},{"name":"sahasra","email":"ram","Password":"ravan"}]

    @staticmethod
    def getTestData(test_case_name):
        dict = {}
        book = openpyxl.load_workbook(r"C:\Users\DEEPA\PythonTesting\test_data.xlsx")
        sheet = book.active
        for i in range(1,sheet.max_row + 1):
            if sheet.cell(row=i,column=1).value == test_case_name:
                for j in range(2,sheet.max_column + 1):
                    dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

        return[dict]